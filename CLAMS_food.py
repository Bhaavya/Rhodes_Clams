import csv
import sys
import datetime
import matplotlib
import matplotlib.pyplot as plt
import math
import os
import openpyxl





def main(argv,ofile,day_start,night_start):
 ofile=ofile.replace('"','')
 if os.path.isfile(ofile):
    wb=openpyxl.load_workbook(ofile)
 else:
    wb=openpyxl.Workbook()
 sheet = wb.get_sheet_names()
 l=len(sheet)

 if l>1:

   if sheet[0]=='CLAMS Respiratory':
       wb.remove_sheet(wb.worksheets[1])
       ws=wb.create_sheet()
       ws.title='CLAMS food intake'

   else:
     wb.remove_sheet(wb.worksheets[0])
     ws=wb.create_sheet(0)
     ws.title='CLAMS food intake'
 elif l==1:
     
        if sheet[0]=='CLAMS Respiratory':
          ws=wb.create_sheet()
          ws.title='CLAMS food intake'
        else:
            if sheet[0]=='CLAMS food intake':
                wb.remove_sheet(wb.worksheets[0])
                ws=wb.create_sheet()
                ws.title='CLAMS food intake'
            
            else:
                ws=wb.worksheets[0]
                ws.title='CLAMS food intake'
 r=1

 min_index= day_start.find(':')
 sec_index =day_start.find(':',min_index+1)
 d_min= int(day_start[min_index+1:sec_index])
 d_hr=int(day_start[:min_index])
            
 if ((day_start.lower()).find('am')>0):

    ampm_index=(day_start.lower()).find('am')
    ampm='am'
        
 else :
  if ((day_start.lower()).find ('pm')>0):
    ampm_index=(day_start.lower()).find('pm')
    ampm='pm'
 if (ampm=='am'  and d_hr==12):
    d_hr=0
 else:
  if (ampm=='pm'  and d_hr!=12):
    d_hr+=12
 d_sec= int(day_start[sec_index+1:ampm_index-1])

 min_index_nt= night_start.find(':')
 sec_index_nt =night_start.find(':',min_index_nt+1)
 n_min= int(night_start[min_index_nt+1:sec_index_nt])
 n_hr=int(night_start[:min_index_nt])
                
 if ((night_start.lower()).find('am') >0):
  ampm_index_nt=night_start.find('am')
  ampm_nt='am'
 else :
            
     if ((night_start.lower()).find('pm') >0):
         ampm_index_nt=night_start.find('pm')
         ampm_nt='pm'

 if ampm_nt.lower()=='am'and n_hr==12:
                        n_hr=0
 else:
     if ampm_nt.lower()=='pm' and n_hr!=12:
         n_hr+=12
 n_sec= int(night_start[sec_index_nt+1:ampm_index_nt-1])


 for ifile in argv:


 



  ifile=ifile.replace('"','')
  file1= open (ifile,"rU")
 
 



  read=csv.reader(file1,delimiter=',',dialect=csv.excel_tab)
  row=1
  while row<17:
   next(read)
   row=row+1;
  row1= next(read)
  file1.close()


  file1.close()
  file1= open (ifile,"rU")
  read=csv.reader(file1,delimiter=',',dialect=csv.excel_tab)
  row=1
  while row<17:
     next(read)
     row=row+1;
 
  day1=row1[3]
  end_time_prev=''
  meals_am=list()
  meals_pm=list()
  meals=list()
  meals_amd=list()
  mealsd=list()
  meals_pmd=list()
  bouts_time=list()
  i=0
  j=0
  p=0
  k=0
  ap=0
  food=0
  flag=0
  bt=0
  ie=0
  iea=0
  iep=0
  food_total=list()
  food_am=list()
  food_pm=list()
  bouts=list()
  bouts_food=list()
  bouts_am=list()
  bouts_am_index1=list()
  bouts_am_index2=list()
  bouts_pm_index1=list()
  bouts_pm_index2=list()
  bouts_pm=list()
  b1=0
  b2=0
  bouts_am_food=list()
  bouts_pm_food=list()
  bouts_am_dur=list()
  bouts_pm_dur=list()
  nt_start=list()
  nt_end=list()
  max=0.0
  min_time=datetime.datetime(1990,01,01,00,00,00)
  max_time=datetime.datetime(1990,01,01,00,00,00)

  day_index= day1.find('/')
  yr_index =day1.find('/',day_index+1)
  day= int(day1[day_index+1:yr_index])
  mon=int(day1[:day_index])

  yr=int(day1[yr_index+1:])
  new_dt=datetime.datetime(yr,mon,day,d_hr,d_min,d_sec)-datetime.timedelta(days=1)
  yr=new_dt.year
  mon=new_dt.month
  day=new_dt.day


  day_nt= day
  mon_nt=mon

  yr_nt=yr




  if datetime.datetime(yr,mon,day,d_hr,d_min,d_sec)>datetime.datetime(yr_nt,mon_nt,day_nt,n_hr,n_min,n_sec):
   new_dt=datetime.datetime(yr_nt,mon_nt,day_nt,n_hr,n_min,n_sec)+datetime.timedelta(days=1)
   yr_nt=new_dt.year
   mon_nt=new_dt.month
   day_nt=new_dt.day
  intermeal_interval=list()
  intermeal_interval_am=list()
  intermeal_interval_pm=list()
  intrameal_intervals_am=list()
  intrameal_intervals_pm=list()
  sum_bouts=0.0
  meal_dur=list()
  meal_dur_am=list()
  meal_dur_pm=list()
  intrameal_intervals=list()
  prev_nt=datetime.datetime(yr_nt,mon_nt,day_nt,n_hr,n_min,n_sec)-datetime.timedelta(days=1)
  nt_start.append(prev_nt)
  nt_end.append(datetime.datetime(yr,mon,day,d_hr,d_min,d_sec))
  
  for row in read:
  
    
    
   
   if ((row[1]!='')and(float(row[7])>=0.03)):
     beg_mon_cur=row[3]
    
     day_index2= beg_mon_cur.find('/')
     yr_index2 =beg_mon_cur.find('/',day_index2+1)
     day2= int(beg_mon_cur[day_index2+1:yr_index2])
     mon2=int(beg_mon_cur[:day_index2])
     yr2=int(beg_mon_cur[yr_index2+1:])
     beg_time_cur=row[4]
     min_index2= beg_time_cur.find(':')
     sec_index2 =beg_time_cur.find(':',min_index2+1)
     min2= int(beg_time_cur[min_index2+1:sec_index2])
     hr2=int(beg_time_cur[:min_index2])
     if ((beg_time_cur.lower()).find('am') >0):
        ampm2_index=(beg_time_cur.lower()).find('am')
        ampm2='am'
     else :
        ampm2_index=(beg_time_cur.lower()).find('pm')
        ampm2='pm'
     sec2= int(beg_time_cur[sec_index2+1:ampm2_index-1])
     hr2t=hr2
     if ampm2=='am'and hr2==12:
        hr2t=0
     else:
        if ampm2=='pm' and hr2!=12:
            hr2t+=12
     if datetime.datetime(yr2,mon2,day2,hr2t,min2,sec2)> datetime.datetime(yr_nt,mon_nt,day_nt,n_hr,n_min,n_sec):
      mon=mon2
      day=day2
      yr=yr2
      if datetime.datetime(yr,mon,day,d_hr,d_min,d_sec)>datetime.datetime(yr_nt,mon_nt,day_nt,n_hr,n_min,n_sec):
        new_dt=datetime.datetime(yr_nt,mon_nt,day_nt,n_hr,n_min,n_sec)+datetime.timedelta(days=1)
        yr_nt=new_dt.year
        mon_nt=new_dt.month
        day_nt=new_dt.day
      prev_nt=datetime.datetime(yr_nt,mon_nt,day_nt,n_hr,n_min,n_sec)-datetime.timedelta(days=1)
      nt_start.append(prev_nt)
      nt_end.append(datetime.datetime(yr,mon,day,d_hr,d_min,d_sec))
  
     end_time_cur=row[5]
     min_index3= end_time_cur.find(':')
     sec_index3 =end_time_cur.find(':',min_index3+1)
     min3= int(end_time_cur[min_index3+1:sec_index3])
     hr3=int(end_time_cur[:min_index3])
     if ((end_time_cur.lower()).find('am') >0):
        ampm3_index=(end_time_cur.lower()).find('am')
        ampm3='am'
     else :
        ampm3_index=(end_time_cur.lower()).find('pm')
        ampm3='pm'
     sec3= int(end_time_cur[sec_index3+1:ampm3_index-1])
     hr3t=hr3
     if ampm3=='am'and hr3==12:
        hr3t=0
     else:
        if ampm3=='pm' and hr3!=12:
            hr3t+=12
     if hr2t> hr3t:
        newdt=datetime.datetime(yr2,mon2,day2)+datetime.timedelta(days=1)
        yr3=newdt.year
        mon3=newdt.month
        day3=newdt.day
     else:
        yr3=yr2
        mon3=mon2
        day3=day2


     hr_diff1=hr3-hr2
     if (hr_diff1<0):
        hr_diff1=hr_diff1+12
     min_diff1=min3-min2
     if((min_diff1)<0):
            min_diff1=60+min_diff1
            hr_diff1=hr_diff1-1
     sec_diff1=sec3-sec2
     if((sec_diff1)<0):
        sec_diff1=60+sec_diff1
        min_diff1=min_diff1-1
     bout_dur= hr_diff1*3600+ min_diff1*60+sec_diff1
     bout_dur_min=(float)(bout_dur)/60
     if  (datetime.datetime(yr2,mon2,day2,hr2t,min2,sec2)< min_time):
                    min_time=datetime.datetime(yr2,mon2,day2,hr2t,min2,sec2)
                                                        
 
     if datetime.datetime(yr3,mon3,day3,hr3t,min3,sec3)> max_time:
                                  max_time=datetime.datetime(yr3,mon3,day3,hr3t,min3,sec3)
                                  
                                  

     if(end_time_prev!=''):
          bd=datetime.datetime(yr2,mon2,day2,hr2t,min2,sec2)-datetime.datetime(yr1,mon1,day1,hr1t,min1,sec1)
          
          
          bout_diff= (bd.days*1440+ (float)(bd.seconds)/60 )

          
          if(bout_diff>10.0):
              
              meals[i].append((hr1,min1,sec1,ampm1))
              mealsd[i].append((yr1,mon1,day1,hr1t,min1,sec1))
              intrameal_interval=datetime.datetime(mealsd[i][1][0],mealsd[i][1][1],mealsd[i][1][2],mealsd[i][1][3],mealsd[i][1][4],mealsd[i][1][5])-datetime.datetime(mealsd[i][0][0],mealsd[i][0][1],mealsd[i][0][2],mealsd[i][0][3],mealsd[i][0][4],mealsd[i][0][5])
              intrameal_interval_min= intrameal_interval.days*1440+ (float)(intrameal_interval.seconds)/60
              intrameal_intervals.append(intrameal_interval_min)
              
              meals.append([])
              mealsd.append([])
              sum_bouts=(float)(sum_bouts/60)
              meal_dur.append(sum_bouts)
              
              i=i+1
              bouts.append([])
             
              meals[i].append((hr2,min2,sec2,ampm1))
              mealsd[i].append((yr2,mon2,day,hr2t,min2,sec2))
              intermeal_interval.append(bout_diff)
              
              
              
              food_total.append(food)
             
              
              
            
              if ap==0:
              
                meals_am[j].append((hr1,min1,sec1,ampm1))
                meals_amd[j].append((yr1,mon1,day1,hr1t,min1,sec1))
                meal_dur_am.append(sum_bouts)
             
                sum_bouts=0.0
                intermeal_interval_am.append(bout_diff)
                
                intrameal_intervals_am.append(intrameal_interval_min)
               
                food_am.append(food)
                food=0
                j=j+1
              else:
                meals_pm[k].append((hr1,min1,sec1,ampm1))
                meals_pmd[k].append((yr1,mon1,day1,hr1t,min1,sec1))
                meal_dur_pm.append(sum_bouts)
              
                sum_bouts=0.0
                food_pm.append(food)
                food=0
                intermeal_interval_pm.append(bout_diff)
                
                intrameal_intervals_pm.append(intrameal_interval_min)
               
                k=k+1
                    
              
              if datetime.datetime(yr2,mon2,day2,hr2t,min2,sec2)>=datetime.datetime(yr,mon,day,d_hr,d_min,d_sec) and datetime.datetime(yr2,mon2,day2,hr2t,min2,sec2)<datetime.datetime(yr_nt,mon_nt,day_nt,n_hr,n_min,n_sec):
                  meals_am.append([])
                  meals_amd.append([])
                  meals_am[j].append((hr2,min2,sec2,ampm2))
                  meals_amd[j].append((yr2,mon2,day2,hr2t,min2,sec2))
                  ap=0
              else:
                  
                 meals_pm.append([])
                 meals_pmd.append([])
                 meals_pm[k].append((hr2,min2,sec2,ampm2))
                 meals_pmd[k].append((yr2,mon2,day2,hr2t,min2,sec2))
                
                 ap=1
                     
          if ap==0:
                bouts_am.append([])
                bouts_am[b1].append((yr2,mon2,day2,hr2t,min2,sec2))
                bouts_am[b1].append((yr3,mon3,day3,hr3t,min3,sec3))
                b1=b1+1
                bouts_am_food.append(float(row[7]))
                bouts_am_dur.append(bout_dur_min)
                bouts_am_index1.append(beg_time_cur)
                bouts_am_index2.append(end_time_cur)
                if (float(row[7])>max):
                        max=float(row[7])
          else:
                                    bouts_pm.append([])
                                    bouts_pm[b2].append((yr2,mon2,day2,hr2t,min2,sec2))
                                    bouts_pm[b2].append((yr3,mon3,day3,hr3t,min3,sec3))
                                    b2=b2+1
                                    bouts_pm_food.append(float(row[7]))
                                    bouts_pm_dur.append(bout_dur_min)
                                    bouts_pm_index1.append(beg_time_cur)
                                    bouts_pm_index2.append(end_time_cur)
                                    if (float(row[7])>max):
                                            max=float(row[7])

     else:
       meals.append([]);
       mealsd.append([])
       meals[i].append((hr2,min2,sec2,ampm2))
       mealsd[i].append((yr2,mon2,day2,hr2t,min2,sec2))
       bouts.append([])
       min_time=datetime.datetime(yr2,mon2,day2,hr2t,min2,sec2)
       max_time=datetime.datetime(yr2,mon3,day3,hr3t,min3,sec3)
       if datetime.datetime(yr2,mon2,day2,hr2t,min2,sec2)>=datetime.datetime(yr,mon,day,d_hr,d_min,d_sec) and datetime.datetime(yr2,mon2,day2,hr2t,min2,sec2)<datetime.datetime(yr_nt,mon_nt,day_nt,n_hr,n_min,n_sec):
        meals_am.append([]);
        meals_amd.append([]);
        meals_am[j].append((hr2,min2,sec2,ampm2))
        meals_amd[j].append((yr2,mon2,day2,hr2t,min2,sec2))
        bouts_am.append([])
        bouts_am[b1].append((yr2,mon2,day2,hr2t,min2,sec2))
        bouts_am[b1].append((yr3,mon3,day3,hr3t,min3,sec3))
        b1=b1+1
        bouts_am_index1.append(beg_time_cur)
        bouts_am_index2.append(end_time_cur)
        bouts_am_food.append(float(row[7]))
        bouts_am_dur.append(bout_dur_min)
        if (float(row[7])>max):
            max=float(row[7])
                                
                                
       
       else:
          meals_pm.append([]);
          meals_pmd.append([])
          meals_pm[k].append((hr2,min2,sec2,ampm2))
          meals_pmd[k].append((yr2,mon2,day2,hr2t,min2,sec2))
          bouts_pm.append([])
          bouts_pm[b2].append((yr2,mon2,day2,hr2t,min2,sec2))
          bouts_pm[b2].append((yr3,mon3,day3,hr3t,min3,sec3))
          b2=b2+1
          bouts_pm_food.append(float(row[7]))
          bouts_pm_index1.append(beg_time_cur)
          bouts_pm_index2.append(end_time_cur)
          bouts_pm_dur.append(bout_dur_min)
          if (float(row[7])>max):
                  max=float(row[7])
          ap=1

     sum_bouts+=bout_dur
     food+=float(row[7])
     bouts[i].append(beg_time_cur)
     bouts[i].append(end_time_cur)
     bouts_food.append(float(row[7]))
    

    
     bouts_time.append(bout_dur_min)
    

     end_time_prev= row[5]


    
     min_index1= end_time_prev.find(':')
    
     sec_index1= end_time_prev.find(':',min_index1+1)
    
     min1= int(end_time_prev[min_index1+1:sec_index1])
    
     hr1=int(end_time_prev[:min_index1])
    
     if ((end_time_prev.lower()).find('am') >0):
        ampm1_index=(end_time_prev.lower()).find('am')
        ampm1='am'
     else:
        ampm1_index=(end_time_prev.lower()).find('pm')
        ampm1='pm'
     hr1t=hr1
     if ampm1=='am'and hr1==12:
        hr1t=0
     else:
        if ampm1=='pm' and hr1!=12:
                    hr1t+=12
                        
     sec1= int(end_time_prev[sec_index1+1:ampm1_index-1])
     if hr2t> hr1t:
      newdt=datetime.datetime(yr2,mon2,day2)+datetime.timedelta(days=1)
      yr1=newdt.year
      mon1=newdt.month
      day1=newdt.day
     else:
      yr1=yr2
      mon1=mon2
      day1=day2

  meals[i].append((hr1,min1,sec1,ampm1))
  mealsd[i].append((yr1,mon1,day1,hr1t,min1,sec1))
  hr_diff2=meals[i][1][0]-meals[i][0][0]
  intrameal_interval=datetime.datetime(mealsd[i][1][0],mealsd[i][1][1],mealsd[i][1][2],mealsd[i][1][3],mealsd[i][1][4],mealsd[i][1][5])-datetime.datetime(mealsd[i][0][0],mealsd[i][0][1],mealsd[i][0][2],mealsd[i][0][3],mealsd[i][0][4],mealsd[i][0][5])
  intrameal_interval_min= intrameal_interval.days*1440+ (float)(intrameal_interval.seconds)/60

  sum_bouts=float(sum_bouts/60);
             
  intrameal_intervals.append(intrameal_interval_min)

  meal_dur.append(sum_bouts)
 
 

  food_total.append(food)
  if ap==0:
    
    meals_am[j].append((hr1,min1,sec1,ampm1))
    meals_amd[j].append((yr1,mon1,day1,hr1t,min1,sec1))
    meal_dur_am.append(sum_bouts)
    food_am.append(food)
    food=0
    sum_bouts=0
   
    intrameal_intervals_am.append(intrameal_interval_min)
   
    j=j+1

  else:
            meals_pm[k].append((hr1,min1,sec1,ampm1))
            meals_pmd[k].append((yr1,mon1,day1,hr1t,min1,sec1))
            meal_dur_pm.append(sum_bouts)
          

            sum_bouts=0.0
            food_pm.append(food)
            food=0
            intrameal_intervals_pm.append(intrameal_interval_min)
           

            k=k+1
 

  def add_zero(s):
   if s<10:
    return '0'+str(s)
   else:
     return str(s)
  if meals !=[]:
    i=0
    
    
    #writer.writerow(['Output for  '+ifile])
    ws.cell(row=r,column=1,value='Output for  '+ifile)
    r=r+1
    #writer.writerow (["All Meals :"])
    ws.cell(row=r,column=1,value="All Meals :")
    r=r+1
    #writer.writerow(["Count","From","To","Food Consumption (g)","Meal Duration (min) ","Intrameal interval (min) ","Intermeal interval(min)"])
    Title=["Count","From","To","Food Consumption (g)","Meal Duration (min) ","Intrameal interval (min) ","Intermeal interval(min)"]
    for c in range(1,len(Title)+1):
      ws.cell(row=r,column=c,value=Title[c-1])
    t=0
    r=r+1
    for meal in meals:
        
            if i>0:
                  temp=intermeal_interval[i-1]
            else:
              temp=' '
            #writer.writerow([i+1,str(meal[0][0])+':'+str(meal[0][1])+':'+str(meal[0][2])+' '+str(meal[0][3]),str(meal[1][0])+':'+str(meal[1][1])+':'+str(meal[1][2])+' '+str(meal[1][3]),food_total[i],str(meal_dur[i]),str(intrameal_intervals[i]),temp])
            Values=[i+1,str(meal[0][0])+':'+add_zero(meal[0][1])+':'+add_zero(meal[0][2])+' '+str(meal[0][3]),str(meal[1][0])+':'+add_zero(meal[1][1])+':'+add_zero(meal[1][2])+' '+str(meal[1][3]),food_total[i],meal_dur[i],intrameal_intervals[i],temp]
            for c in range(1,len(Values)+1):
              ws.cell(row=r,column=c,value=Values[c-1])
            r=r+1
            #writer.writerow ([' ','Bouts: '])
            #writer.writerow([' ','From','To','Food (g)','Duration(min)'])
            ws.cell(row=r,column=2,value='Bouts: ')
            r=r+1
            Values=[' ','From','To','Food (g)','Duration(min)']
            for c in range(1,len(Values)+1):
                ws.cell(row=r,column=c,value=Values[c-1])
            r=r+1
            j=0
            k=1
            
            for b in bouts[i]:
                if j%2==0:
              
                   
                   #writer.writerow([k+')',b,bouts[i][j+1],bouts_food[t],str(bouts_time[t])])
                   Values=[str(k)+')',b,bouts[i][j+1],bouts_food[t],bouts_time[t]]
                  
                   for c in range(1,len(Values)+1):
                     ws.cell(row=r,column=c,value=Values[c-1])
                   r=r+1
                   t=t+1
                   k=k+1
                j=j+1
                
           
        
        
            #writer.writerow([' '])
            r=r+1
            i=i+1



  i=0
 #writer.writerow([' '])
  r=r+2

  if meals_am !=[]:
   #writer.writerow(["Meals during light cycle:"])
   ws.cell(row=r,column=1,value="Meals during light cycle:")
   r=r+2
   #writer.writerow([' '])
   #writer.writerow(['Count','From ','To','Food (g)','Meal Duration(min)','Intrameal interval(min)','Intermeal interval(min)'])
   Title=["Count","From","To","Food Consumption (g)","Meal Duration (min) ","Intrameal interval (min) ","Intermeal interval(min)"]
   for c in range(1,len(Title)+1):
      ws.cell(row=r,column=c,value=Title[c-1])
   r=r+1
   for meal in meals_am:
    if i>0:
           temp=intermeal_interval_am[i-1]
    else:
       temp= ' '
   #writer.writerow([i+1,str(meal[0][0])+':'+str(meal[0][1])+':'+str(meal[0][2])+' '+str(meal[0][3]),str(meal[1][0])+':'+str(meal[1][1])+':'+str(meal[1][2])+' '+str(meal[1][3]),food_am[i],str(meal_dur_am[i]),str(intrameal_intervals_am[i]),temp])
    Values=[i+1,str(meal[0][0])+':'+add_zero(meal[0][1])+':'+add_zero(meal[0][2])+' '+str(meal[0][3]),str(meal[1][0])+':'+add_zero(meal[1][1])+':'+add_zero(meal[1][2])+' '+str(meal[1][3]),food_am[i],meal_dur_am[i],intrameal_intervals_am[i],temp]
    for c in range(1,len(Values)+1):
           ws.cell(row=r,column=c,value=Values[c-1])
    i=i+1
    r=r+1

  i=0
  #writer.writerow([' '])
  r=r+2
  if meals_pm !=[]:
  #writer.writerow(["Meals during dark cycle:"])
  
   ws.cell(row=r,column=1,value="Meals during dark cycle:")
  #writer.writerow([' '])
   r=r+2
  #writer.writerow(['Count','From','To','Food (g)','Meal Duration(min)','Intrameal interval(min)','Intermeal interval(min)'])
   Title=["Count","From","To","Food Consumption (g)","Meal Duration (min) ","Intrameal interval (min) ","Intermeal interval(min)"]
   for c in range(1,len(Title)+1):
          ws.cell(row=r,column=c,value=Title[c-1])
   r=r+1
   for meal in meals_pm:
     if i>0:
         temp=intermeal_interval_pm[i-1]
     else:
        temp= ' '
          #writer.writerow([i+1,str(meal[0][0])+':'+str(meal[0][1])+':'+str(meal[0][2])+' '+str(meal[0][3]),str(meal[1][0])+':'+str(meal[1][1])+':'+str(meal[1][2])+' '+str(meal[1][3]),food_pm[i],str(meal_dur_pm[i]),str(intrameal_intervals_pm[i]),temp])
     Values=[i+1,str(meal[0][0])+':'+add_zero(meal[0][1])+':'+add_zero(meal[0][2])+' '+str(meal[0][3]),str(meal[1][0])+':'+add_zero(meal[1][1])+':'+add_zero(meal[1][2])+' '+str(meal[1][3]),food_pm[i],meal_dur_pm[i],intrameal_intervals_pm[i],temp]
     for c in range(1,len(Values)+1):
            ws.cell(row=r,column=c,value=Values[c-1])
     r=r+1
     i=i+1
  r=r+2
  file1.close()
 wb.save(ofile)
if __name__=='__main__':
    ofile=raw_input('Enter output filename \t')
    ofile=ofile.strip()
    day_start=raw_input('Enter start time for light cycle as hr:min:sec am/pm \t')
    night_start=raw_input('Enter start time for dark cycle as hr:min:sec am/pm \t')
    main(sys.argv[1:],ofile,day_start,night_start)


 








