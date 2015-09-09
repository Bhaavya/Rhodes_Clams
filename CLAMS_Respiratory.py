import csv
import sys
import datetime
import os
import openpyxl
def main(ifiles,ofile,night_start,day_start,interval):
 ofile=ofile.strip()
 min_index= day_start.find(':')
 sec_index =day_start.find(':',min_index+1)
 d_min= int(day_start[min_index+1:sec_index])
 d_hr=int(day_start[:min_index])
 r=1
 if ((day_start.lower()).find('am')>0):

    ampm_index=(day_start.lower()).find('am')
    ampm='am'
        
 else :
  if ((day_start.lower()).find ('pm')>0):
    ampm_index=(day_start.lower()).find('pm')
    ampm='pm'
 if (ampm=='am'  and d_hr==12):
    d_hr=0
 else:
  if (ampm=='pm'  and d_hr!=12):
    d_hr+=12
 d_sec= int(day_start[sec_index+1:ampm_index-1])

 min_index_nt= night_start.find(':')
 sec_index_nt =night_start.find(':',min_index_nt+1)
 n_min= int(night_start[min_index_nt+1:sec_index_nt])
 n_hr=int(night_start[:min_index_nt])
 prev_hr=n_hr
 if ((night_start.lower()).find('am') >0):
  ampm_index_nt=night_start.find('am')
  ampm_nt='am'
 else :
            
     if ((night_start.lower()).find('pm') >0):
         ampm_index_nt=night_start.find('pm')
         ampm_nt='pm'
                        
 if ampm_nt.lower()=='am'and n_hr==12:
                        n_hr=0
 else:
    if ampm_nt.lower()=='pm' and n_hr!=12:
        n_hr+=12
 n_sec= int(night_start[sec_index_nt+1:ampm_index_nt-1])
 ofile=ofile.replace('"','')
 if os.path.isfile(ofile):
  wb=openpyxl.load_workbook(ofile)
 else:
  wb=openpyxl.Workbook()

 sheet = wb.get_sheet_names()
 l=len(sheet)
 if l>1:

  if sheet[0]=='CLAMS food intake':
      wb.remove_sheet(wb.worksheets[1])
      ws=wb.create_sheet()
      ws.title='CLAMS Respiratory'
  else:
      wb.remove_sheet(wb.worksheets[0])
      ws=wb.create_sheet(0)
      ws.title='CLAMS food intake'
 else:
     if l==1:
         if sheet[0]=='CLAMS food intake':
            ws=wb.create_sheet()
            ws.title='CLAMS Respiratory'
         else:
             if sheet[0]=='CLAMS Respiratory':
                 wb.remove_sheet(wb.worksheets[0])
                 ws=wb.create_sheet()
                 ws.title='CLAMS Respiratory'
             
             else:
                     ws=wb.worksheets[0]
                     ws.title='CLAMS Respiratory'

 for ifile in ifiles:
  
  ifile=ifile.replace('"','')
  file1= open (ifile,"rU")
  read=csv.reader(file1,delimiter=',',dialect=csv.excel_tab)
  row=1
  while row<26:
   row=row+1;
   next(read)
  row1= next(read)
  day1=row1[2]
  file1.close()
  file1= open (ifile,"rU")
  read=csv.reader(file1,delimiter=',',dialect=csv.excel_tab)
  row=1
  while row<26:
     next(read)
     row=row+1;
  day_beg_index=day1.find(' ')
  day=day1[:day_beg_index]
  day_index= day1.find('/')
  yr_index =day1.find('/',day_index+1)
  day= int(day1[day_index+1:yr_index])
  mon=int(day1[:day_index])

  yr=int(day1[yr_index+1:day_beg_index])
  end_time_prev=''



  prev_min= n_min
  prev_hrt=n_hr
  prev_sec=n_sec
 

  pampm=ampm_nt


  prev_time=datetime.datetime(yr,mon,day,prev_hrt,prev_min,prev_sec)

     
  int_time=row1[2][day_beg_index+1:]
  min_index3= int_time.find(':')
  sec_index3=int_time.find(':',min_index3+1)
  min3= int(int_time[min_index3+1:sec_index3])
  hr3=int(int_time[:min_index3])
  if ((int_time.lower()).find('am') >0):
             ampm3_index=(int_time.lower()).find('am')
             ampm3='am'
  else :
                  ampm3_index=(int_time.lower()).find('pm')
                  ampm3='pm'
  sec3= int(int_time[sec_index3+1:ampm3_index-1])
  hr3t=hr3
  if ampm3=='am'and hr3==12:
                     hr3t=0
  else:
        if ampm3=='pm' and hr3!=12:
          hr3t+=12
  yr3=yr
  mon3=mon
  day3=day

  new_dt=datetime.datetime(yr,mon,day,d_hr,d_min,d_sec)-datetime.timedelta(days=1)
  yr=new_dt.year
  mon=new_dt.month
  day=new_dt.day
                
  day_nt= day
  mon_nt=mon
  yr_nt=yr

  if datetime.datetime(yr,mon,day,d_hr,d_min,d_sec)>datetime.datetime(yr_nt,mon_nt,day_nt,n_hr,n_min,n_sec):
                new_dt=datetime.datetime(yr_nt,mon_nt,day_nt,n_hr,n_min,n_sec)+datetime.timedelta(days=1)
                yr_nt=new_dt.year
                mon_nt=new_dt.month
                day_nt=new_dt.day

  t=0
    
    
  slots=list()
  intervals=list()
            
  slots_am=list()
  slots_pm=list()
                    
                    
  avg_rer=0.0
  sum_rer=0.0
  avg_o2=0.0
  sum_o2=0.0
  avg_co2=0.0
  sum_co2=0.0
  sum_heat=0.0
  avg_heat=0.0
  if datetime.datetime(yr3,mon3,day3,hr3t,min3,sec3)< prev_time:
   prev_time=prev_time-datetime.timedelta(days=1)
  prev_time_2=prev_time+datetime.timedelta(minutes=interval)
  while prev_time_2 <= datetime.datetime(yr3,mon3,day3,hr3t,min3,sec3):
    prev_time=prev_time+datetime.timedelta(minutes=interval)

    prev_time_2=prev_time+datetime.timedelta(minutes=interval)
    pampm='am'
    hour1=prev_time.hour
    minute1=prev_time.minute
    second1=prev_time.second
    if minute1<10:
        prev_mins='0'+str(minute1)
    else:
        prev_mins=str(minute1)
    if second1<10:
     prev_secs='0'+str(second1)
    else:
        prev_secs=str(second1)
    if hour1==12:
        pampm='pm'
    if hour1>12:
        hour1=hour1-12
        pampm='pm'
    if hour1==0:
        hour1=12
    yr1=prev_time.year
    mon1=prev_time.month
    day1=prev_time.day
    hr1t=prev_time.hour
    hr1=hour1
    min1=prev_time.minute
    sec1=prev_time.second


    slots.append((str(hour1),prev_mins,prev_secs,pampm,0.0,0.0,0.0,0.0,[]))
    
    
    yr1=prev_time.year
    mon1=prev_time.month
    day1=prev_time.day
    hr1t=prev_time.hour
    hr1=hour1
    min1=prev_time.minute
    sec1=prev_time.second
    
    
    if datetime.datetime(yr1,mon1,day1,hr1t,min1,sec1)>datetime.datetime(yr,mon,day,d_hr,d_min,d_sec) and datetime.datetime(yr1,mon1,day1,hr1t,min1,sec1)<=datetime.datetime(yr_nt,mon_nt,day_nt,n_hr,n_min,n_sec):
        
        slots_am.append((str(hour1),prev_mins,prev_secs,pampm,0.0,0.0,0.0,0.0,[]))
    
    else:
        slots_pm.append((str(hour1),prev_mins,prev_secs,pampm,0.0,0.0,0.0,0.0,[]))

                    


  for row in read:
 
  
   if len(row)>1:
    if (float)(row[4])!=0.0:
      
      beg_mon_cur_index= row[2].find(' ')
      beg_mon_cur=row[2][:beg_mon_cur_index]
      day_index2= beg_mon_cur.find('/')
      yr_index2 =beg_mon_cur.find('/',day_index2+1)
      
      day2= int(beg_mon_cur[day_index2+1:yr_index2])
      mon2=int(beg_mon_cur[:day_index2])
      yr2=int(beg_mon_cur[yr_index2+1:])
     
      beg_time_cur=row[2][beg_mon_cur_index+1:]
      min_index2= beg_time_cur.find(':')
      sec_index2 =beg_time_cur.find(':',min_index2+1)
      min2= int(beg_time_cur[min_index2+1:sec_index2])
      hr2=int(beg_time_cur[:min_index2])
      if ((beg_time_cur.lower()).find('am') >0):
          ampm2_index=(beg_time_cur.lower()).find('am')
          ampm2='am'
      else :
            ampm2_index=(beg_time_cur.lower()).find('pm')
            ampm2='pm'
      sec2= int(beg_time_cur[sec_index2+1:ampm2_index-1])
      hr2t=hr2
      if ampm2=='am'and hr2==12:
                    hr2t=0
      else:
            if ampm2=='pm' and hr2!=12:
                    hr2t+=12
                                                          

      if datetime.datetime(yr2,mon2,day2,hr2t,min2,sec2)> datetime.datetime(yr_nt,mon_nt,day_nt,n_hr,n_min,n_sec):
        mon=mon2
        day=day2
        yr=yr2
        if datetime.datetime(yr,mon,day,d_hr,d_min,d_sec)>datetime.datetime(yr_nt,mon_nt,day_nt,n_hr,n_min,n_sec):
          new_dt=datetime.datetime(yr_nt,mon_nt,day_nt,n_hr,n_min,n_sec)+datetime.timedelta(days=1)
          yr_nt=new_dt.year
          mon_nt=new_dt.month
          day_nt=new_dt.day
          
   
   
      if (datetime.datetime(yr2,mon2,day2,hr2t,min2,sec2)>prev_time_2):
     
     
  
       
        
        its=intervals
        avg_rer=sum_rer/t
        avg_o2=sum_o2/t
        avg_co2=sum_co2/t
        avg_heat=sum_heat/t
        sum_rer=0.0
        sum_co2=0.0
        sum_o2=0.0
        sum_heat=0.0
        t=0
        intervals=list()
        
        


        prev_time=prev_time+datetime.timedelta(minutes=interval)

        prev_time_2=prev_time+datetime.timedelta(minutes=interval)
        pampm='am'
        hour1=prev_time.hour
        minute1=prev_time.minute
        second1=prev_time.second
        if minute1<10:
            prev_mins='0'+str(minute1)
        else:
           prev_mins=str(minute1)
        if second1<10:
            prev_secs='0'+str(second1)
        else:
            prev_secs=str(second1)
        if hour1==12:
            pampm='pm'
        if hour1>12:
            hour1=hour1-12
            pampm='pm'
        if hour1==0:
            hour1=12
        slots.append((str(hour1),prev_mins,prev_secs,pampm,avg_o2,avg_co2,avg_rer,avg_heat,its))
        
     

        
        yr1=prev_time.year
        mon1=prev_time.month
        day1=prev_time.day
        hr1t=prev_time.hour
        hr1=hour1
        min1=prev_time.minute
        sec1=prev_time.second
        if datetime.datetime(yr1,mon1,day1,hr1t,min1,sec1)>datetime.datetime(yr,mon,day,d_hr,d_min,d_sec) and datetime.datetime(yr1,mon1,day1,hr1t,min1,sec1)<=datetime.datetime(yr_nt,mon_nt,day_nt,n_hr,n_min,n_sec):
            
              slots_am.append((str(hour1),prev_mins,prev_secs,pampm,avg_o2,avg_co2,avg_rer,avg_heat,its))
        
        else:
            slots_pm.append((str(hour1),prev_mins,prev_secs,pampm,avg_o2,avg_co2,avg_rer,avg_heat,its))
          




      
      
      
      
      
    
      sum_rer=sum_rer+(float)(row[13])
      sum_co2=sum_co2+(float)(row[8])
      sum_o2=sum_o2+(float)(row[3])
      sum_heat=sum_heat+(float)(row[14])
      t=t+1
      intervals.append(((yr2,mon2,day2),beg_time_cur,row[3],row[8],row[13],row[14]))
 

     

  its=intervals
  avg_rer=sum_rer/t
  avg_o2=sum_o2/t
  avg_co2=sum_co2/t
  avg_heat=sum_heat/t
  sum_rer=0.0
  sum_co2=0.0
  sum_o2=0.0
  sum_heat=0.0
  t=0
  intervals=list()
                                                                                                         


  prev_time=prev_time+datetime.timedelta(minutes=interval)
  prev_time_2=prev_time+datetime.timedelta(minutes=interval)
  pampm='am'
  hour1=prev_time.hour
  minute1=prev_time.minute
  second1=prev_time.second
  if minute1<10:
            prev_mins='0'+str(minute1)
  else:
        prev_mins=str(minute1)
  if second1<10:
            prev_secs='0'+str(second1)
  else:
            prev_secs=str(second1)
  if hour1==12:
    pampm='pm'
  if hour1>12:
            hour1=hour1-12
            pampm='pm'
  if hour1==0:
            hour1=12
        
        
  slots.append((str(hour1),prev_mins,prev_secs,pampm,avg_o2,avg_co2,avg_rer,avg_heat,its))


  yr1=prev_time.year
  mon1=prev_time.month
  day1=prev_time.day
  hr1t=prev_time.hour
  hr1=hour1
  min1=prev_time.minute
  sec1=prev_time.second

    
  if datetime.datetime(yr1,mon1,day1,hr1t,min1,sec1)>datetime.datetime(yr,mon,day,d_hr,d_min,d_sec) and datetime.datetime(yr1,mon1,day1,hr1t,min1,sec1)<=datetime.datetime(yr_nt,mon_nt,day_nt,n_hr,n_min,n_sec):
            
            slots_am.append((str(hour1),prev_mins,prev_secs,pampm,avg_o2,avg_co2,avg_rer,avg_heat,its))

  else:
      slots_pm.append((str(hour1),prev_mins,prev_secs,pampm,avg_o2,avg_co2,avg_rer,avg_heat,its))

  ynt=yr_nt
  mnt=mon_nt
  dnt=day_nt
  while prev_time <= datetime.datetime(ynt,mnt,dnt,n_hr,n_min,n_sec):
    prev_time=prev_time+datetime.timedelta(minutes=interval)
    pampm='am'
    hour1=prev_time.hour
    minute1=prev_time.minute
    second1=prev_time.second
    if minute1<10:
            prev_mins='0'+str(minute1)
    else:
        prev_mins=str(minute1)
    if second1<10:
            prev_secs='0'+str(second1)
    else:
                        prev_secs=str(second1)
    if hour1==12:
            pampm='pm'
    if hour1>12:
            hour1=hour1-12
            pampm='pm'
    if hour1==0:
            hour1=12
    yr1=prev_time.year
    mon1=prev_time.month
    day1=prev_time.day
    hr1t=prev_time.hour
    hr1=hour1
    min1=prev_time.minute
    sec1=prev_time.second
    if datetime.datetime(yr1,mon1,day1,hr1t,min1,sec1)> datetime.datetime(yr_nt,mon_nt,day_nt,n_hr,n_min,n_sec):
             mon=mon2
             day=day2
             yr=yr2
             if datetime.datetime(yr,mon,day,d_hr,d_min,d_sec)>datetime.datetime(yr_nt,mon_nt,day_nt,n_hr,n_min,n_sec):
                         new_dt=datetime.datetime(yr_nt,mon_nt,day_nt,n_hr,n_min,n_sec)+datetime.timedelta(days=1)
                         yr_nt=new_dt.year
                         mon_nt=new_dt.month
                         day_nt=new_dt.day

    slots.append((str(hour1),prev_mins,prev_secs,pampm,0.0,0.0,0.0,0.0,[]))
    
    
    yr1=prev_time.year
    mon1=prev_time.month
    day1=prev_time.day
    hr1t=prev_time.hour
    hr1=hour1
    min1=prev_time.minute
    sec1=prev_time.second

                            
    if datetime.datetime(yr1,mon1,day1,hr1t,min1,sec1)>datetime.datetime(yr,mon,day,d_hr,d_min,d_sec) and datetime.datetime(yr1,mon1,day1,hr1t,min1,sec1)<=datetime.datetime(yr_nt,mon_nt,day_nt,n_hr,n_min,n_sec):

          slots_am.append((str(hour1),prev_mins,prev_secs,pampm,0.0,0.0,0.0,0.0,[]))
    
    else:
        slots_pm.append((str(hour1),prev_mins,prev_secs,pampm,0.0,0.0,0.0,0.0,[]))

  if slots!=[]:
      ws.cell(row=r,column=1,value='Output for'+ifile)
      r=r+2
      Title=['Count','Time','Average VO2 (ml/kg/hr)','Average VCO2(ml/kg/hr)','Average RER','Average Heat(kcal/hr)']
      for c in range(1,len(Title)+1):
        ws.cell(row=r,column=c,value=Title[c-1])
      r=r+1
      i=0
      for s in slots:
        Values=[i+1,str(s[0])+':'+str(s[1])+':'+str(s[2])+' '+str(s[3]),s[4],s[5],s[6],s[7]]
        for c in range(1,len(Values)+1):

            ws.cell(row=r,column=c,value=Values[c-1])
        r=r+1
        if s[8]!=[]:
         ws.cell(row=r,column=2,value='Intervals: ')
         r=r+1
         Title=['Count','Time','V02 (ml/kg/hr)','VCO2 (ml/kg/hr)','RER','Heat (kcal/hr)']
        
         for c in range(1,len(Title)+1):
           ws.cell(row=r,column=c,value=Title[c-1])
         r=r+1
         k='a'
        
         for b in s[8]:
            Values=[k+')',str(b[1]),float(b[2]),float(b[3]),float(b[4]),float(b[5])]
            for c in range(1,len(Values)+1):
            
              ws.cell(row=r,column=c,value=Values[c-1])
            k=chr(ord(k)+1)
            r=r+1
        i=i+1
        r=r+1
  if slots_am!=[]:
    r=r+1
    ws.cell(row=r,column=1,value='Averages during Light Cycle: ')
    r=r+2
    Title=['Count','Time','Average VO2 (ml/kg/hr)','Average VCO2(ml/kg/hr)','Average RER','Average Heat(kcal/hr)']
    for c in range(1,len(Title)+1):
        ws.cell(row=r,column=c,value=Title[c-1])
    r=r+1
    i=0
    for s in slots_am:
        Values=[i+1,str(s[0])+':'+str(s[1])+':'+str(s[2])+' '+str(s[3]),s[4],s[5],s[6],s[7]]
        for c in range(1,len(Values)+1):
            
            ws.cell(row=r,column=c,value=Values[c-1])
        r=r+1
        i=i+1
  if slots_pm!=[]:
        r=r+1
        ws.cell(row=r,column=1,value='Averages during Dark Cycle: ')
        r=r+2
        Title=['Count','Time','Average VO2 (ml/kg/hr)','Average VCO2(ml/kg/hr)','Average RER','Average Heat(kcal/hr)']
        for c in range(1,len(Title)+1):
           ws.cell(row=r,column=c,value=Title[c-1])
        r=r+1
        i=0
        for s in slots_pm:
          Values=[i+1,str(s[0])+':'+str(s[1])+':'+str(s[2])+' '+str(s[3]),s[4],s[5],s[6],s[7]]
          for c in range(1,len(Values)+1):
            
            ws.cell(row=r,column=c,value=Values[c-1])
          r=r+1
          i=i+1
  r=r+2
  file1.close()
 wb.save(ofile)
if __name__=='__main__':
 ofile=raw_input('Enter output filename \t')

 day_start=raw_input('Enter start time for light cycle as hr:min:sec am/pm \t')
 night_start=raw_input('Enter start time for dark cycle as hr:min:sec am/pm \t')
 interval= (float)(raw_input('Enter time interval in minutes for calculating average  \t'))
 main(sys.argv[1:],ofile,night_start,day_start,interval)
